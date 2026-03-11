"""

This is the file we created for read and write a excel file
means we write data in excel and we read data in excel


"""
import openpyxl

class excel_method:

    """This class is created for to read and write data in excel file """

    @staticmethod   # This is the method which can be called by class name only . There is no need of creation of object
    def get_rows(excel_path ,sheet_name):

        excel_file = openpyxl.load_workbook(excel_path)
        sheet = excel_file[sheet_name]
        return sheet.max_row



    # if u want to read the excel file . so what you need
    #### 1.Excel file
    #### 2.sheet (because single excel file contains multiple sheets)
    #### 3.row    (which row want to read)
    #### 4.column (which column want to read)

    @staticmethod
    def read_excel(excel_path , sheet_name , row , column):

        excel_file = openpyxl.load_workbook(excel_path)
        sheet = excel_file[sheet_name]
        data = sheet.cell(row , column).value  #cell is the function who have row and column argument
        return data

    @staticmethod
    def write_excel(excel_path , sheet_name , row , column , data):

        excel_file = openpyxl.load_workbook(excel_path)
        sheet = excel_file[sheet_name]
        sheet.cell(row , column , data)  # write data in cell
        excel_file.save(excel_path) # filename.save()
        excel_file.close()






